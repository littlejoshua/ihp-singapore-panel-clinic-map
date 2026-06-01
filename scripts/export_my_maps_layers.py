#!/usr/bin/env python3
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs"
INPUT_CSV = OUTPUT_DIR / "google_matched_all.csv"
LAYER_DIR = OUTPUT_DIR / "my_maps_layers"

CATEGORIES = ["GP", "Specialist", "TCM"]
BANDS = ["4.5+", "4.0-4.4", "<4.0 / no rating"]

LAYER_FILENAMES = {
    ("GP", "4.5+"): "01_GP_4.5_plus.csv",
    ("GP", "4.0-4.4"): "02_GP_4.0_to_4.4.csv",
    ("GP", "<4.0 / no rating"): "03_GP_below_4_or_no_rating.csv",
    ("Specialist", "4.5+"): "04_Specialist_4.5_plus.csv",
    ("Specialist", "4.0-4.4"): "05_Specialist_4.0_to_4.4.csv",
    ("Specialist", "<4.0 / no rating"): "06_Specialist_below_4_or_no_rating.csv",
    ("TCM", "4.5+"): "07_TCM_4.5_plus.csv",
    ("TCM", "4.0-4.4"): "08_TCM_4.0_to_4.4.csv",
    ("TCM", "<4.0 / no rating"): "09_TCM_below_4_or_no_rating.csv",
}

MAP_FIELDS = [
    "clinic_name",
    "category",
    "rating_band",
    "address",
    "phone",
    "specialty",
    "google_place_name",
    "google_rating",
    "google_review_count",
    "open_in_google_maps",
    "latitude",
    "longitude",
    "match_status",
    "review_priority",
    "match_confidence",
    "postal_code",
    "google_postal_code",
    "google_formatted_address",
    "match_notes",
]


def read_rows(path):
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path, rows, fields):
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path, rows, fields):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Review"
    sheet.append(fields)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        sheet.append([row.get(field, "") for field in fields])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        max_length = len(str(column_cells[0].value or ""))
        for cell in column_cells[1:101]:
            max_length = max(max_length, len(str(cell.value or "")))
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
            max(max_length + 2, 10),
            48,
        )
    workbook.save(path)


def map_row(row):
    mapped = dict(row)
    mapped["open_in_google_maps"] = row.get("google_maps_url", "")
    return mapped


def main():
    LAYER_DIR.mkdir(exist_ok=True)
    rows = [map_row(row) for row in read_rows(INPUT_CSV)]

    summary_rows = []
    for category in CATEGORIES:
        for band in BANDS:
            layer_rows = [
                row
                for row in rows
                if row.get("category") == category
                and row.get("rating_band") == band
                and row.get("latitude")
                and row.get("longitude")
            ]
            filename = LAYER_FILENAMES[(category, band)]
            write_csv(LAYER_DIR / filename, layer_rows, MAP_FIELDS)
            summary_rows.append(
                {
                    "layer_file": filename,
                    "category": category,
                    "rating_band": band,
                    "record_count": len(layer_rows),
                    "high_priority_count": sum(
                        1 for row in layer_rows if row.get("review_priority") == "high"
                    ),
                    "needs_review_count": sum(
                        1 for row in layer_rows if row.get("match_status") == "needs_review"
                    ),
                }
            )

    review_rows = [
        row
        for row in rows
        if row.get("review_priority") == "high"
        or row.get("match_status") in {"needs_review", "not_found", "api_error"}
    ]
    write_csv(OUTPUT_DIR / "google_matched_high_priority_review.csv", review_rows, MAP_FIELDS)
    write_xlsx(OUTPUT_DIR / "google_matched_high_priority_review.xlsx", review_rows, MAP_FIELDS)
    write_csv(
        LAYER_DIR / "00_layer_summary.csv",
        summary_rows,
        [
            "layer_file",
            "category",
            "rating_band",
            "record_count",
            "high_priority_count",
            "needs_review_count",
        ],
    )

    instructions = """Google My Maps import steps

1. Open https://www.google.com/mymaps and create a new map.
2. Import each numbered CSV in this folder as one layer.
3. When asked for location columns, choose latitude and longitude.
4. When asked for marker title, choose clinic_name.
5. Rename each layer from its CSV file name, for example GP - 4.5+.
6. Marker details include open_in_google_maps, which links to the original Google Maps place.
"""
    (LAYER_DIR / "README_import_steps.txt").write_text(instructions, encoding="utf-8")

    print(f"Wrote {len(summary_rows)} layer CSVs to {LAYER_DIR}")
    for row in summary_rows:
        print(row)
    print(f"High priority review rows: {len(review_rows)}")


if __name__ == "__main__":
    main()
