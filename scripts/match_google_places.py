#!/usr/bin/env python3
import argparse
import csv
import difflib
import json
import os
import re
import ssl
import time
import urllib.error
import urllib.request
from pathlib import Path

import certifi
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs"
INPUT_CSV = OUTPUT_DIR / "ihp_clinics_master_pending_google.csv"
CACHE_PATH = OUTPUT_DIR / "google_places_cache.json"
TEXT_SEARCH_URL = "https://places.googleapis.com/v1/places:searchText"

TEXT_SEARCH_FIELDS = ",".join(
    [
        "places.id",
        "places.displayName",
        "places.formattedAddress",
        "places.location",
        "places.rating",
        "places.userRatingCount",
        "places.googleMapsUri",
        "places.types",
    ]
)
SSL_CONTEXT = ssl.create_default_context(cafile=certifi.where())

OUTPUT_FIELDS = [
    "source_file",
    "category",
    "s_n",
    "ihp_clinic_id",
    "region",
    "area",
    "clinic_name",
    "address",
    "postal_code",
    "phone",
    "specialty",
    "remarks",
    "phpc",
    "search_query",
    "google_place_id",
    "google_place_name",
    "google_formatted_address",
    "google_rating",
    "google_review_count",
    "google_maps_url",
    "latitude",
    "longitude",
    "match_status",
    "match_confidence",
    "name_similarity",
    "address_similarity",
    "postal_match",
    "google_postal_code",
    "matched_query",
    "google_types",
    "review_priority",
    "match_notes",
    "rating_band",
    "layer_name",
]


def read_api_key():
    key = os.environ.get("GOOGLE_MAPS_API_KEY", "").strip()
    if key:
        return key

    env_path = ROOT / ".env"
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            name, value = line.split("=", 1)
            if name.strip() == "GOOGLE_MAPS_API_KEY":
                return value.strip().strip('"').strip("'")

    key_path = ROOT / ".google_maps_api_key"
    if key_path.exists():
        return key_path.read_text(encoding="utf-8").strip()

    raise SystemExit(
        "Missing GOOGLE_MAPS_API_KEY. Set it as an environment variable, "
        "or put GOOGLE_MAPS_API_KEY=... in .env."
    )


def load_rows(path):
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path, rows, fields=OUTPUT_FIELDS):
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def load_cache():
    if not CACHE_PATH.exists():
        return {}
    return json.loads(CACHE_PATH.read_text(encoding="utf-8"))


def save_cache(cache):
    CACHE_PATH.write_text(
        json.dumps(cache, indent=2, sort_keys=True, ensure_ascii=False),
        encoding="utf-8",
    )


def normalize(value):
    value = str(value or "").upper()
    value = re.sub(r"[^A-Z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def extract_postal_code(value):
    match = re.search(r"(?<!\d)(?:S(?:INGAPORE)?\s*)?(\d{6})(?!\d)", str(value or ""), re.I)
    return match.group(1) if match else ""


def rating_band(rating):
    if rating in ("", None):
        return "<4.0 / no rating"
    try:
        value = float(rating)
    except (TypeError, ValueError):
        return "<4.0 / no rating"
    if value >= 4.5:
        return "4.5+"
    if value >= 4.0:
        return "4.0-4.4"
    return "<4.0 / no rating"


def layer_name(category, band):
    return f"{category} - {band}"


def text_search(api_key, query):
    payload = {
        "textQuery": query,
        "regionCode": "SG",
        "languageCode": "en",
        "maxResultCount": 3,
    }
    request = urllib.request.Request(
        TEXT_SEARCH_URL,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "X-Goog-Api-Key": api_key,
            "X-Goog-FieldMask": TEXT_SEARCH_FIELDS,
        },
        method="POST",
    )

    try:
        with urllib.request.urlopen(request, timeout=30, context=SSL_CONTEXT) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as error:
        detail = error.read().decode("utf-8", errors="replace")
        return {"error": {"status_code": error.code, "body": detail}}


def best_candidate(row, places, query):
    if not places:
        return None, 0, "not_found", {
            "name_similarity": "",
            "address_similarity": "",
            "postal_match": "",
            "matched_query": query,
            "google_types": "",
            "match_notes": "No Google Places candidates returned.",
        }

    expected_name = normalize(row["clinic_name"])
    expected_postal = row.get("postal_code") or extract_postal_code(row.get("address"))
    expected_address = normalize(row.get("address"))

    scored = []
    for place in places:
        place_name = place.get("displayName", {}).get("text", "")
        place_address = place.get("formattedAddress", "")
        place_postal = extract_postal_code(place_address)
        candidate_text = normalize(f"{place_name} {place_address}")

        name_ratio = difflib.SequenceMatcher(None, expected_name, normalize(place_name)).ratio()
        address_ratio = difflib.SequenceMatcher(None, expected_address, normalize(place_address)).ratio()
        postal_match = bool(expected_postal and place_postal and expected_postal == place_postal)
        postal_mismatch = bool(expected_postal and place_postal and expected_postal != place_postal)
        type_match = bool({"doctor", "health", "hospital"} & set(place.get("types", [])))

        score = 0.62 * name_ratio + 0.23 * address_ratio
        if postal_match:
            score += 0.1
        if postal_mismatch:
            score -= 0.2
        if type_match:
            score += 0.05
        if expected_name and expected_name in candidate_text:
            score += 0.05
        scored.append((max(min(score, 1.0), 0), place, postal_match, postal_mismatch, place_postal, name_ratio, address_ratio))

    scored.sort(key=lambda item: item[0], reverse=True)
    score, place, postal_match, postal_mismatch, place_postal, name_ratio, address_ratio = scored[0]

    if postal_mismatch:
        status = "needs_review"
    elif score >= 0.78 and (postal_match or name_ratio >= 0.72):
        status = "matched"
    elif score >= 0.62 and name_ratio >= 0.65:
        status = "name_match_address_partial"
    elif postal_match and address_ratio >= 0.55:
        status = "address_only"
    else:
        status = "needs_review"

    notes = []
    if postal_match:
        notes.append("postal code matches")
    elif postal_mismatch:
        notes.append("postal code differs")
    else:
        notes.append("postal code not confirmed")
    if name_ratio < 0.65:
        notes.append("clinic name differs")
    elif name_ratio < 0.82:
        notes.append("clinic name partially matches")
    else:
        notes.append("clinic name closely matches")
    if address_ratio < 0.55:
        notes.append("address differs")
    elif address_ratio < 0.75:
        notes.append("address partially matches")
    else:
        notes.append("address closely matches")

    diagnostics = {
        "name_similarity": round(name_ratio, 3),
        "address_similarity": round(address_ratio, 3),
        "postal_match": "yes" if postal_match else "no",
        "google_postal_code": place_postal,
        "matched_query": query,
        "google_types": "; ".join(place.get("types", [])),
        "review_priority": "high" if postal_mismatch else ("medium" if status != "matched" else "low"),
        "match_notes": "; ".join(notes),
    }
    return place, round(score, 3), status, diagnostics


def apply_match(row, place, confidence, status, diagnostics=None):
    diagnostics = diagnostics or {}
    if not place:
        row.update(
            {
                "google_place_id": "",
                "google_place_name": "",
                "google_formatted_address": "",
                "google_rating": "",
                "google_review_count": "",
                "google_maps_url": "",
                "latitude": "",
                "longitude": "",
                "match_status": status,
                "match_confidence": confidence,
                "name_similarity": diagnostics.get("name_similarity", ""),
                "address_similarity": diagnostics.get("address_similarity", ""),
                "postal_match": diagnostics.get("postal_match", ""),
                "google_postal_code": diagnostics.get("google_postal_code", ""),
                "matched_query": diagnostics.get("matched_query", ""),
                "google_types": diagnostics.get("google_types", ""),
                "review_priority": diagnostics.get("review_priority", ""),
                "match_notes": diagnostics.get("match_notes", ""),
            }
        )
        return row

    location = place.get("location", {})
    rating = place.get("rating", "")
    band = rating_band(rating)
    row.update(
        {
            "google_place_id": place.get("id", ""),
            "google_place_name": place.get("displayName", {}).get("text", ""),
            "google_formatted_address": place.get("formattedAddress", ""),
            "google_rating": rating,
            "google_review_count": place.get("userRatingCount", ""),
            "google_maps_url": place.get("googleMapsUri", ""),
            "latitude": location.get("latitude", ""),
            "longitude": location.get("longitude", ""),
            "match_status": status,
            "match_confidence": confidence,
            "name_similarity": diagnostics.get("name_similarity", ""),
            "address_similarity": diagnostics.get("address_similarity", ""),
            "postal_match": diagnostics.get("postal_match", ""),
            "google_postal_code": diagnostics.get("google_postal_code", ""),
            "matched_query": diagnostics.get("matched_query", ""),
            "google_types": diagnostics.get("google_types", ""),
            "review_priority": diagnostics.get("review_priority", ""),
            "match_notes": diagnostics.get("match_notes", ""),
            "rating_band": band,
            "layer_name": layer_name(row["category"], band),
        }
    )
    return row


def query_variants(row):
    variants = [
        row.get("search_query", ""),
        f"{row.get('clinic_name', '')} {row.get('postal_code', '')} Singapore",
        f"{row.get('clinic_name', '')} Singapore",
    ]
    seen = set()
    for query in variants:
        query = re.sub(r"\s+", " ", query).strip()
        if query and query not in seen:
            seen.add(query)
            yield query


def match_row(row, api_key, cache, sleep_seconds):
    best_place = None
    best_confidence = 0
    best_status = "not_found"
    best_diagnostics = {}

    for query in query_variants(row):
        if query not in cache:
            cache[query] = text_search(api_key, query)
            save_cache(cache)
            if sleep_seconds:
                time.sleep(sleep_seconds)

        result = cache[query]
        if "error" in result:
            row["match_status"] = "api_error"
            row["match_confidence"] = ""
            row["remarks"] = f"{row.get('remarks', '')} API error: {result['error']}".strip()
            return row

        place, confidence, status, diagnostics = best_candidate(row, result.get("places", []), query)
        if confidence > best_confidence:
            best_place = place
            best_confidence = confidence
            best_status = status
            best_diagnostics = diagnostics
        if status in {"matched", "name_match_address_partial"}:
            return apply_match(row, place, confidence, status, diagnostics)

    return apply_match(row, best_place, best_confidence, best_status, best_diagnostics)


def write_xlsx(path, rows):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["field", "value"])
    for cell in summary[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    status_counts = {}
    for row in rows:
        status_counts[row.get("match_status", "")] = status_counts.get(row.get("match_status", ""), 0) + 1
    summary.append(["total", len(rows)])
    for status, count in sorted(status_counts.items()):
        summary.append([status, count])
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 14

    def add_sheet(title, sheet_rows):
        sheet = workbook.create_sheet(title)
        sheet.append(OUTPUT_FIELDS)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for row in sheet_rows:
            sheet.append([row.get(field, "") for field in OUTPUT_FIELDS])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            max_length = len(str(column_cells[0].value or ""))
            for cell in column_cells[1:101]:
                max_length = max(max_length, len(str(cell.value or "")))
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 10), 48)

    add_sheet("All Matches", rows)
    add_sheet(
        "Manual Review",
        [
            row
            for row in rows
            if row.get("match_status")
            in {"name_match_address_partial", "address_only", "needs_review", "not_found", "api_error"}
        ],
    )
    add_sheet("Matched", [row for row in rows if row.get("match_status") == "matched"])
    workbook.save(path)


def main():
    parser = argparse.ArgumentParser(description="Match IHP clinics to Google Places.")
    parser.add_argument("--limit", type=int, default=20, help="Number of rows to process.")
    parser.add_argument("--offset", type=int, default=0, help="Starting row offset.")
    parser.add_argument("--all", action="store_true", help="Process all rows.")
    parser.add_argument("--sleep", type=float, default=0.05, help="Seconds to sleep after uncached API calls.")
    args = parser.parse_args()

    api_key = read_api_key()
    rows = load_rows(INPUT_CSV)
    selected = rows[args.offset :] if args.all else rows[args.offset : args.offset + args.limit]
    cache = load_cache()

    matched = []
    for index, row in enumerate(selected, start=args.offset + 1):
        print(f"[{index}/{len(rows)}] {row['category']} | {row['clinic_name']}")
        matched.append(match_row(row, api_key, cache, args.sleep))

    suffix = "all" if args.all else f"sample_{args.offset + 1}_{args.offset + len(selected)}"
    output_path = OUTPUT_DIR / f"google_matched_{suffix}.csv"
    write_csv(output_path, matched)
    xlsx_path = OUTPUT_DIR / f"google_matched_{suffix}.xlsx"
    write_xlsx(xlsx_path, matched)

    review_rows = [
        row for row in matched if row.get("match_status") in {"needs_review", "not_found", "api_error"}
    ]
    write_csv(OUTPUT_DIR / f"needs_review_{suffix}.csv", review_rows)

    counts = {}
    for row in matched:
        counts[row.get("match_status", "")] = counts.get(row.get("match_status", ""), 0) + 1
    print("Status counts:", counts)
    print(f"Wrote {output_path}")
    print(f"Wrote {xlsx_path}")


if __name__ == "__main__":
    main()
