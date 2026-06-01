#!/usr/bin/env python3
import csv
import html
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs"

NS = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

FILES = [
    {
        "path": ROOT / "QBE PanelClinicsSingapore GP_As of October 2025.xlsx",
        "category": "GP",
    },
    {
        "path": ROOT / "IHP SP Panel Clinic Listing October 2025.xlsx",
        "category": "Specialist",
    },
    {
        "path": ROOT / "PanelClinics TCM_As of October 2025.xlsx",
        "category": "TCM",
    },
]

MASTER_FIELDS = [
    "source_file",
    "category",
    "s_n",
    "ihp_clinic_id",
    "region",
    "area",
    "clinic_name",
    "address",
    "postal_code",
    "phone",
    "specialty",
    "remarks",
    "phpc",
    "search_query",
    "google_place_id",
    "google_place_name",
    "google_formatted_address",
    "google_rating",
    "google_review_count",
    "google_maps_url",
    "latitude",
    "longitude",
    "match_status",
    "match_confidence",
    "rating_band",
    "layer_name",
]


def column_index(cell_ref):
    match = re.match(r"([A-Z]+)", cell_ref)
    if not match:
        return 0

    index = 0
    for char in match.group(1):
        index = index * 26 + ord(char) - ord("A") + 1
    return index - 1


def read_shared_strings(workbook_zip):
    try:
        root = ET.fromstring(workbook_zip.read("xl/sharedStrings.xml"))
    except KeyError:
        return []

    strings = []
    for item in root.findall("a:si", NS):
        strings.append("".join(text.text or "" for text in item.findall(".//a:t", NS)))
    return strings


def cell_value(cell, shared_strings):
    cell_type = cell.attrib.get("t")

    if cell_type == "s":
        value = cell.find("a:v", NS)
        if value is None or value.text is None:
            return ""
        return shared_strings[int(value.text)]

    if cell_type == "inlineStr":
        return "".join(text.text or "" for text in cell.findall(".//a:t", NS))

    value = cell.find("a:v", NS)
    return value.text if value is not None and value.text is not None else ""


def read_first_sheet(path):
    with zipfile.ZipFile(path) as workbook_zip:
        shared_strings = read_shared_strings(workbook_zip)
        sheet = ET.fromstring(workbook_zip.read("xl/worksheets/sheet1.xml"))

    rows = []
    for row in sheet.findall(".//a:sheetData/a:row", NS):
        values = []
        for cell in row.findall("a:c", NS):
            index = column_index(cell.attrib["r"])
            while len(values) <= index:
                values.append("")
            values[index] = cell_value(cell, shared_strings)
        rows.append(values)
    return rows


def clean_text(value):
    value = html.unescape(str(value or ""))
    value = value.replace("\xa0", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def clean_phone(value):
    return clean_text(value)


def extract_postal_code(address, explicit_postal_code):
    explicit = clean_text(explicit_postal_code)
    if re.fullmatch(r"\d{6}", explicit):
        return explicit

    match = re.search(r"(?<!\d)(?:S(?:INGAPORE)?\s*)?(\d{6})(?!\d)", address, re.I)
    return match.group(1) if match else ""


def normalize_header(value):
    return re.sub(r"\s+", " ", clean_text(value)).upper()


def find_header_row(rows):
    for index, row in enumerate(rows):
        headers = [normalize_header(value) for value in row]
        if "CLINIC NAME" in headers and "ADDRESS" in headers:
            return index, headers
    raise ValueError("Could not find clinic header row")


def row_value(row, header_map, header_name):
    index = header_map.get(header_name)
    if index is None or index >= len(row):
        return ""
    return clean_text(row[index])


def rating_band(rating):
    if rating == "":
        return "<4.0 / no rating"
    try:
        value = float(rating)
    except ValueError:
        return "<4.0 / no rating"
    if value >= 4.5:
        return "4.5+"
    if value >= 4.0:
        return "4.0-4.4"
    return "<4.0 / no rating"


def layer_name(category, band):
    return f"{category} - {band}"


def parse_workbook(config):
    rows = read_first_sheet(config["path"])
    header_index, headers = find_header_row(rows)
    header_map = {header: index for index, header in enumerate(headers)}

    records = []
    for row in rows[header_index + 1 :]:
        clinic_name = row_value(row, header_map, "CLINIC NAME")
        address = row_value(row, header_map, "ADDRESS")
        if not clinic_name or not address:
            continue

        postal_code = extract_postal_code(
            address,
            row_value(row, header_map, "POSTAL CODE"),
        )
        category = config["category"]
        band = rating_band("")

        record = {
            "source_file": config["path"].name,
            "category": category,
            "s_n": row_value(row, header_map, "S/N"),
            "ihp_clinic_id": row_value(row, header_map, "IHP CLINIC ID"),
            "region": row_value(row, header_map, "REGION"),
            "area": row_value(row, header_map, "AREA"),
            "clinic_name": clinic_name,
            "address": address,
            "postal_code": postal_code,
            "phone": clean_phone(
                row_value(row, header_map, "TEL NO.")
                or row_value(row, header_map, "TEL NO")
            ),
            "specialty": row_value(row, header_map, "SPECIALTY"),
            "remarks": row_value(row, header_map, "REMARKS"),
            "phpc": row_value(row, header_map, "PHPC"),
            "search_query": clean_text(f"{clinic_name} {address} Singapore"),
            "google_place_id": "",
            "google_place_name": "",
            "google_formatted_address": "",
            "google_rating": "",
            "google_review_count": "",
            "google_maps_url": "",
            "latitude": "",
            "longitude": "",
            "match_status": "pending_google_match",
            "match_confidence": "",
            "rating_band": band,
            "layer_name": layer_name(category, band),
        }
        records.append(record)

    return records


def write_csv(path, rows, fields=MASTER_FIELDS):
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_sheet(workbook, title, rows, fields):
    sheet = workbook.create_sheet(title)
    sheet.append(fields)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in rows:
        sheet.append([row.get(field, "") for field in fields])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        header = str(column_cells[0].value or "")
        max_length = len(header)
        for cell in column_cells[1:101]:
            max_length = max(max_length, len(str(cell.value or "")))
        width = min(max(max_length + 2, 10), 42)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def write_xlsx(path, records):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["category", "record_count"])
    for cell in summary[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for category in ["GP", "Specialist", "TCM"]:
        count = sum(1 for row in records if row["category"] == category)
        summary.append([category, count])
    summary.append(["Total", len(records)])
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 14

    write_sheet(workbook, "Master Pending Google", records, MASTER_FIELDS)
    write_sheet(
        workbook,
        "Google Matching Input",
        records,
        [
            "category",
            "clinic_name",
            "address",
            "postal_code",
            "phone",
            "specialty",
            "search_query",
        ],
    )
    for category in ["GP", "Specialist", "TCM"]:
        category_records = [row for row in records if row["category"] == category]
        write_sheet(workbook, category, category_records, MASTER_FIELDS)

    workbook.save(path)


def data_quality_issues(records):
    issues = []
    for row in records:
        base = {
            "category": row["category"],
            "clinic_name": row["clinic_name"],
            "address": row["address"],
            "source_file": row["source_file"],
            "search_query": row["search_query"],
        }
        if not row["postal_code"]:
            issues.append({**base, "issue": "missing_postal_code"})
        if not row["phone"]:
            issues.append({**base, "issue": "missing_phone"})
    return issues


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)

    records = []
    for config in FILES:
        records.extend(parse_workbook(config))

    write_csv(OUTPUT_DIR / "ihp_clinics_master_pending_google.csv", records)
    write_csv(
        OUTPUT_DIR / "google_matching_input.csv",
        records,
        [
            "category",
            "clinic_name",
            "address",
            "postal_code",
            "phone",
            "specialty",
            "search_query",
        ],
    )

    for category in ["GP", "Specialist", "TCM"]:
        category_records = [row for row in records if row["category"] == category]
        write_csv(OUTPUT_DIR / f"{category.lower()}_clinics_pending_google.csv", category_records)

    write_csv(
        OUTPUT_DIR / "data_quality_issues.csv",
        data_quality_issues(records),
        ["issue", "category", "clinic_name", "address", "source_file", "search_query"],
    )
    write_xlsx(OUTPUT_DIR / "ihp_clinics_pending_google.xlsx", records)

    print(f"Wrote {len(records)} records to {OUTPUT_DIR}")
    for category in ["GP", "Specialist", "TCM"]:
        count = sum(1 for row in records if row["category"] == category)
        print(f"{category}: {count}")


if __name__ == "__main__":
    main()
